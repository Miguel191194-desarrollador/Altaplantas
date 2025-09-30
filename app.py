from flask import Flask, render_template, request, redirect, flash, send_file, url_for
from datetime import datetime
from openpyxl import load_workbook
import os, logging, threading, base64, requests

# Carga .env en local; en Render usará sus env vars
try:
    from dotenv import load_dotenv
    load_dotenv()
except Exception:
    pass

app = Flask(__name__)
app.secret_key = 'supersecretkey'
logging.basicConfig(level=logging.INFO)

# ====== Variables desde Render (.env en local solo para pruebas) ======
GAS_WEBHOOK_URL   = os.getenv("GAS_WEBHOOK_URL")                 # URL de tu Apps Script (/exec)
MAIL_TO_ADMIN     = os.getenv("MAIL_TO_ADMIN")                   # opcional, para pruebas
TESORERIA_EMAIL   = os.getenv("TESORERIA_EMAIL", "tesoreria@dimensasl.com")  # copia siempre

SAVE_FOLDER = 'formularios_guardados_plantas'
os.makedirs(SAVE_FOLDER, exist_ok=True)

# ----------------- Rutas -----------------

@app.route('/')
@app.route('/plantas')
@app.route('/formulario_plantas')
def formulario_plantas():
    return render_template('plantas.html')

@app.route('/guardar_plantas', methods=['POST'])
def guardar_plantas():
    data = request.form.to_dict()

    # 1) Generar y guardar Excel en disco
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    nombre_cliente = (data.get('nombre_cliente') or 'cliente').replace(" ", "_")
    filename = f'{nombre_cliente}_{timestamp}.xlsx'
    file_path = os.path.join(SAVE_FOLDER, filename)
    try:
        crear_excel_plantas_solas_a_archivo(data, file_path)
    except Exception:
        logging.exception("❌ Error generando el Excel de plantas")
        flash("Ha ocurrido un error generando el Excel.")
        return redirect(url_for('formulario_plantas'))

    # 2) Enviar correo vía Gmail (Apps Script) en segundo plano
    destinatario = (data.get('correo_comercial') or MAIL_TO_ADMIN or TESORERIA_EMAIL).strip()
    asunto = "Nuevo formulario de alta de plantas (solo plantas)"
    texto  = "Se ha recibido un formulario con alta de plantas."
    html   = "<p>Se ha recibido un formulario con alta de plantas.</p>"

    logging.info(f"📧 Envío principal a: {destinatario}")
    _lanza_envio_seguro(
        to_email=destinatario, subject=asunto, text=texto, html=html, attachment_path=file_path
    )

    # Copia a Tesorería SIEMPRE (evita duplicar si ya es el mismo destino)
    if destinatario.lower() != TESORERIA_EMAIL.lower():
        logging.info(f"📧 Copia a Tesorería: {TESORERIA_EMAIL}")
        _lanza_envio_seguro(
            to_email=TESORERIA_EMAIL, subject=asunto, text=texto, html=html, attachment_path=file_path
        )

    return render_template('gracias.html')

@app.route('/descargar-ultimo-planta')
def descargar_ultimo_excel_planta():
    archivos = [f for f in os.listdir(SAVE_FOLDER) if f.endswith('.xlsx')]
    if not archivos:
        return "No hay archivos de plantas para descargar."
    archivos.sort(reverse=True)
    return send_file(os.path.join(SAVE_FOLDER, archivos[0]), as_attachment=True)

# Debug: comprobar variables en marcha
@app.route('/_env')
def _env():
    ok_url = "OK" if GAS_WEBHOOK_URL else "MISSING"
    ok_admin = "SET" if MAIL_TO_ADMIN else "EMPTY"
    tesor = TESORERIA_EMAIL or "EMPTY"
    return f"GAS_WEBHOOK_URL: {ok_url} | MAIL_TO_ADMIN: {ok_admin} | TESORERIA_EMAIL: {tesor}"

# Test de envío aislado (sin formulario)
@app.route('/_mail_test')
def _mail_test():
    try:
        # Prueba: envía a Admin (si hay) y SIEMPRE copia a Tesorería
        destino_prueba = (MAIL_TO_ADMIN or TESORERIA_EMAIL).strip()
        enviar_via_gmail_webhook(
            to_email=destino_prueba,
            subject="Prueba desde Render (Gmail webhook)",
            text="Hola",
            html="<b>Hola</b>",
            attachment_path=None
        )
        if destino_prueba.lower() != TESORERIA_EMAIL.lower():
            enviar_via_gmail_webhook(
                to_email=TESORERIA_EMAIL,
                subject="Prueba (copia Tesorería)",
                text="Hola Tesorería",
                html="<b>Hola Tesorería</b>",
                attachment_path=None
            )
        return "OK"
    except Exception as e:
        logging.exception("❌ Falló el envío de prueba")
        return f"ERROR: {e}", 500

# -------------- Lógica de Excel --------------

def crear_excel_plantas_solas_a_archivo(data, file_path):
    """
    Plantilla: 'Copia de alta de plantas solas.xlsx'
    Reglas:
      - A2 = Nombre del cliente
      - Las plantas empiezan en la fila 5
      - Orden de columnas: B, D, C, E, F, G, H, I, J, K, L, M
    """
    wb = load_workbook("Copia de alta de plantas solas.xlsx")
    ws = wb.active

    ws["A2"] = data.get("nombre_cliente", "")

    columnas = ["B", "D", "C", "E", "F", "G", "H", "I", "J", "K", "L", "M"]
    campos = [
        "planta_nombre_{}", "planta_direccion_{}", "planta_cp_{}", "planta_poblacion_{}",
        "planta_provincia_{}", "planta_telefono_{}", "planta_email_{}", "planta_horario_{}",
        "planta_observaciones_{}", "planta_contacto_nombre_{}", "planta_contacto_telefono_{}",
        "planta_contacto_email_{}"
    ]

    for i in range(1, 11):
        fila = 3 + i  # empieza en 5
        valores = [data.get(campo.format(i), "") for campo in campos]
        if not (valores[0] or "").strip():
            continue
        for col, val in zip(columnas, valores):
            ws[f"{col}{fila}"] = val

    wb.save(file_path)

# -------------- Envío vía Gmail (Apps Script) --------------

def _lanza_envio_seguro(**kwargs):
    """Lanza el envío en un hilo y captura errores en log."""
    def _seguro():
        try:
            enviar_via_gmail_webhook(**kwargs)
        except Exception:
            logging.exception("❌ Falló el envío de correo (webhook)")
    threading.Thread(target=_seguro, daemon=True).start()

def enviar_via_gmail_webhook(to_email, subject, text, html, attachment_path=None):
    if not GAS_WEBHOOK_URL:
        raise RuntimeError("Falta GAS_WEBHOOK_URL")

    payload = {
        "to": to_email,
        "subject": subject,
        "text": text or "",
        "html": html or (text or "")
    }

    if attachment_path and os.path.exists(attachment_path):
        with open(attachment_path, "rb") as f:
            payload["attachmentBase64"] = base64.b64encode(f.read()).decode("utf-8")
            payload["filename"] = os.path.basename(attachment_path)
            payload["mimeType"] = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    r = requests.post(GAS_WEBHOOK_URL, json=payload, timeout=20)
    if r.status_code != 200 or "OK" not in r.text:
        raise RuntimeError(f"Webhook Gmail error: {r.status_code} {r.text}")
    logging.info(f"✅ Correo enviado vía Gmail (Apps Script) a {to_email}")

# -------------- Main --------------

if __name__ == '__main__':
    port = int(os.environ.get("PORT", 5000))
    app.run(host='0.0.0.0', port=port)





